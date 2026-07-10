import streamlit as st
import pandas as pd
from datetime import datetime, date
import os
import io

# ---------------------------------------------------------------------------
# Optional dependencies — every one of these is wrapped so the app NEVER
# crashes if a package fails to install on the deploy target. The related
# feature just quietly disables itself and shows a small note instead.
# ---------------------------------------------------------------------------
try:
    from num2words import num2words
    NUM2WORDS_AVAILABLE = True
except ImportError:
    NUM2WORDS_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from streamlit_drawable_canvas import st_canvas
    from PIL import Image
    CANVAS_AVAILABLE = True
except ImportError:
    CANVAS_AVAILABLE = False

APP_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(APP_DIR, "assets", "logo.png")

st.set_page_config(page_title="Oromia Bank - Digital Forms", page_icon="🏦", layout="centered")

# ---------------------------------------------------------------------------
# STYLING
# ---------------------------------------------------------------------------
st.markdown("""
<style>
.bank-title { font-size: 24px; font-weight: 800; letter-spacing: 0.5px; }
.bank-sub { font-size: 13px; color: #444; margin-bottom: 4px; }
.section-box {
    border: 1px solid #999; border-radius: 6px; padding: 12px 16px;
    margin-bottom: 14px; background-color: #fafafa;
}
.denom-total { font-weight: 700; font-size: 18px; color: #145214; }
.ai-box {
    border: 1px dashed #6a4fd6; border-radius: 6px; padding: 10px 14px;
    margin-bottom: 14px; background-color: #f5f2ff;
}
.sig-note { font-size: 12px; color: #777; }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# APP HEADER (logo)
# ---------------------------------------------------------------------------
hcol1, hcol2 = st.columns([3, 1])
with hcol1:
    st.markdown('<div class="bank-title">OROMIA BANK — DIGITAL FORMS</div>', unsafe_allow_html=True)
    st.markdown('<div class="bank-sub">Baankii Oromiyaa | ባንኪ ኦሮሚያ — for teller / cashier use</div>', unsafe_allow_html=True)
with hcol2:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, use_container_width=True)
    else:
        st.markdown("**Oromia Bank**")

st.markdown("---")

# ---------------------------------------------------------------------------
# SHARED HELPERS
# ---------------------------------------------------------------------------
ACCOUNT_TYPES = [
    "Saving Account (Herrega Qusannoo)",
    "Special Saving Account (Herrega Qusannoo Addaa)",
    "Demand Account (Herrega Socho'aa)",
    "Other (Specify)",
]

# A small starter list of common Oromia Bank branches for the AI autocomplete helper.
# Tellers can still type any branch name manually — this just speeds up common entries.
KNOWN_BRANCHES = [
    "Gerji", "Bole", "Piassa", "Merkato", "Adama", "Jimma", "Bishoftu", "Sebeta",
    "Shashamane", "Nekemte", "Ambo", "Hawassa", "Dire Dawa", "Harar", "Asella",
]

COMMON_NARRATIVES = [
    "Salary Transfer", "Rent Payment", "Tuition Fee Payment", "Family Support",
    "Business Payment", "Loan Repayment", "Utility Bill Payment", "Other",
]


def ai_format_phone(phone_key: str):
    """AI helper: normalize a typed phone number into +251XXXXXXXXX format."""
    raw = str(st.session_state.get(phone_key, "") or "")
    digits = "".join(ch for ch in raw if ch.isdigit())
    if not digits:
        return
    if digits.startswith("251"):
        digits = digits[3:]
    digits = digits.lstrip("0")
    if len(digits) == 9:
        st.session_state[phone_key] = "+251" + digits
    else:
        st.session_state[phone_key] = raw  # leave as-is if it doesn't look like a valid local number


def ai_smart_check(fields: dict) -> list:
    """AI helper: scans a dict of {label: value} plus special keys and returns a list of
    human-readable issues — missing fields, malformed phone numbers, short account numbers,
    denomination mismatches, etc. Fully local/deterministic, no external calls."""
    issues = []
    for label, value in fields.get("required", {}).items():
        if value is None or str(value).strip() == "" or (isinstance(value, (int, float)) and value <= 0):
            issues.append(f"Missing or empty: **{label}**")

    phone = fields.get("phone")
    if phone:
        digits = "".join(ch for ch in str(phone) if ch.isdigit())
        if len(digits) not in (9, 10, 12, 13):
            issues.append(f"Phone number '{phone}' doesn't look like a valid Ethiopian number.")

    acct_no = fields.get("account_no")
    if acct_no and len(str(acct_no).strip()) < 6:
        issues.append(f"Account number '{acct_no}' looks too short — double-check it.")

    if "denom_total" in fields and "amount" in fields:
        if fields["amount"] and fields["denom_total"] != fields["amount"]:
            issues.append(
                f"Denomination total ({fields['denom_total']:,.2f}) doesn't match the stated amount ({fields['amount']:,})."
            )
    return issues


def check_duplicate(file_name: str, account_no: str, amount, window_minutes: int = 10) -> bool:
    """AI helper: flags likely duplicate submissions — same account number and amount
    submitted again within a short time window (simple fraud/double-entry guard)."""
    if not os.path.exists(file_name) or not account_no:
        return False
    try:
        df = pd.read_csv(file_name)
    except Exception:
        return False
    acct_col = "Account No" if "Account No" in df.columns else "Debit Account No"
    if acct_col not in df.columns or "Timestamp" not in df.columns:
        return False
    amt_col = next((c for c in ["Amount in Number", "Amount (Denomination Total)"] if c in df.columns), None)
    if amt_col is None:
        return False
    recent = df[df[acct_col].astype(str) == str(account_no)]
    recent = recent[recent[amt_col] == amount]
    if recent.empty:
        return False
    try:
        last_time = pd.to_datetime(recent["Timestamp"]).max()
        return (datetime.now() - last_time).total_seconds() < window_minutes * 60
    except Exception:
        return False


def load_demo_data(values: dict):
    st.session_state.update(values)


def clear_form_fields(keys: list):
    for k in keys:
        if k in st.session_state:
            del st.session_state[k]


def _simple_number_to_words(n: int) -> str:
    """Dependency-free fallback used only if num2words isn't installed."""
    ones = ["zero", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
            "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen",
            "seventeen", "eighteen", "nineteen"]
    tens = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]

    def chunk(num):
        if num == 0:
            return ""
        if num < 20:
            return ones[num]
        if num < 100:
            return tens[num // 10] + (" " + ones[num % 10] if num % 10 else "")
        return ones[num // 100] + " hundred" + (" " + chunk(num % 100) if num % 100 else "")

    if n == 0:
        return "zero"
    scales = [(1_000_000_000, "billion"), (1_000_000, "million"), (1_000, "thousand"), (1, "")]
    words, remaining = [], n
    for value, name in scales:
        if remaining >= value:
            count = remaining // value
            remaining %= value
            part = chunk(count)
            words.append(part + (" " + name if name else ""))
    return " ".join(w for w in words if w).strip()


def number_to_words(n: int) -> str:
    return num2words(n).replace("-", " ") if NUM2WORDS_AVAILABLE else _simple_number_to_words(n)


def ai_auto_words(number_key: str, words_key: str):
    amt = int(st.session_state.get(number_key, 0) or 0)
    if amt > 0:
        st.session_state[words_key] = number_to_words(amt)


def ai_auto_denomination(number_key: str, denom_keys: dict):
    """denom_keys maps note value -> session_state key, e.g. {200: 'wd_x200', ...}.
    Optional 'extra' key holds the leftover cents/coins field name."""
    raw_amt = st.session_state.get(number_key, 0) or 0
    amt = int(raw_amt)
    remaining = amt
    for note in [200, 100, 50, 10, 5, 1]:
        if note in denom_keys:
            st.session_state[denom_keys[note]] = int(remaining // note)
            remaining = remaining % note
    if "extra" in denom_keys:
        st.session_state[denom_keys["extra"]] = float(round(raw_amt - amt, 2))


def denomination_block(prefix: str, include_ones=True, extra_label=None, amount_key=None):
    """Renders X200/X100/X50/X10/X5[/X1][/extra_label] number inputs.
    Returns (values_dict, extra_value, total)."""
    n_cols = 5 + (1 if include_ones else 0) + (1 if extra_label else 0)
    cols = st.columns(n_cols)
    values = {}
    i = 0
    for note in [200, 100, 50, 10, 5]:
        with cols[i]:
            values[note] = st.number_input(f"X {note}", min_value=0, step=1, key=f"{prefix}_x{note}")
        i += 1
    if include_ones:
        with cols[i]:
            values[1] = st.number_input("X 1", min_value=0, step=1, key=f"{prefix}_x1")
        i += 1
    extra_val = 0.0
    if extra_label:
        with cols[i]:
            extra_val = st.number_input(extra_label, min_value=0.0, step=0.01, format="%.2f", key=f"{prefix}_extra")

    total = sum(v * k for k, v in values.items()) + extra_val

    denom_keys = {k: f"{prefix}_x{k}" for k in values}
    if extra_label:
        denom_keys["extra"] = f"{prefix}_extra"

    if amount_key:
        b1, b2 = st.columns([1, 1])
        with b1:
            st.button("✨ Auto-write Amount in Words", key=f"{prefix}_btn_words",
                       on_click=ai_auto_words, args=(amount_key, f"{prefix}_words"))
        with b2:
            st.button("✨ Auto-calculate Denomination", key=f"{prefix}_btn_denom",
                       on_click=ai_auto_denomination, args=(amount_key, denom_keys))

    st.markdown(f'<p class="denom-total">Denomination Total: {total:,.2f}</p>', unsafe_allow_html=True)
    return values, extra_val, total


def signature_pad(prefix: str, label: str):
    """Real drawn signature via canvas when available; falls back to a typed name."""
    st.markdown(f"**{label}**")
    if CANVAS_AVAILABLE:
        result = st_canvas(
            fill_color="rgba(255, 255, 255, 0)",
            stroke_width=2,
            stroke_color="#1a1a1a",
            background_color="#ffffff",
            height=130,
            width=320,
            drawing_mode="freedraw",
            key=f"{prefix}_canvas",
        )
        st.markdown('<p class="sig-note">Draw your signature above with mouse/touch/stylus.</p>', unsafe_allow_html=True)
        if result.image_data is not None:
            img = Image.fromarray(result.image_data.astype("uint8"), "RGBA")
            # detect if the canvas is blank (fully transparent)
            if img.getextrema()[3][1] == 0:
                return None
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return buf.getvalue()
        return None
    else:
        st.caption("Drawing pad unavailable on this deployment (install `streamlit-drawable-canvas`) — type your name instead.")
        return st.text_input(f"{label} (typed)", key=f"{prefix}_typed")


def date_field(key: str, label="Date (Guyyaa/Ji'a/Bara)"):
    return st.date_input(label, value=date.today(), key=key, format="DD/MM/YYYY")


def branch_selector(prefix: str, label="Branch / Damee (ቅርንጫፍ)"):
    """AI Assist: quick-pick from known branches, or type a custom one."""
    choice = st.selectbox(
        f"{label} — 🤖 pick or choose 'Type manually'",
        ["Type manually..."] + KNOWN_BRANCHES,
        key=f"{prefix}_branch_choice",
    )
    if choice == "Type manually...":
        return st.text_input(label, key=f"{prefix}_branch")
    st.session_state[f"{prefix}_branch"] = choice
    return choice


def account_type_block(prefix: str):
    st.markdown("**Account Type**")
    acct_type = st.radio("Herrega Qusannoo / የሂሳብ ዓይነት", ACCOUNT_TYPES, key=f"{prefix}_acct_type")
    other = ""
    if acct_type == "Other (Specify)":
        other = st.text_input("Kan Biroo / ሌላ ካልሆነ ይግለጹ (Specify)", key=f"{prefix}_other")
    return acct_type, other


def save_record(file_name: str, record: dict):
    df_new = pd.DataFrame([record])
    if os.path.exists(file_name):
        df_new.to_csv(file_name, mode="a", header=False, index=False)
    else:
        df_new.to_csv(file_name, mode="w", header=True, index=False)


def build_pdf(title: str, record: dict, signature_images: dict) -> bytes:
    """signature_images: {label: png_bytes or None}"""
    if not FPDF_AVAILABLE:
        return b""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Oromia Bank", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, title, ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    for label, value in record.items():
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(58, 7, str(label), border=1)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 7, str(value), border=1, ln=True)

    for label, img_bytes in signature_images.items():
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 7, label, ln=True)
        if img_bytes:
            tmp_path = f"/tmp/_sig_{abs(hash(label))}.png"
            with open(tmp_path, "wb") as f:
                f.write(img_bytes)
            pdf.image(tmp_path, w=60)
            os.remove(tmp_path)
        else:
            pdf.set_font("Helvetica", "I", 9)
            pdf.cell(0, 7, "(no signature captured)", ln=True)

    return bytes(pdf.output(dest="S"))


def export_buttons(prefix: str, record: dict, pdf_title: str, signature_images: dict, key_no: str):
    c1, c2, c3 = st.columns(3)
    with c1:
        csv_bytes = pd.DataFrame([record]).to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ CSV", data=csv_bytes, file_name=f"{prefix}_{key_no}.csv",
                            mime="text/csv", use_container_width=True, key=f"{prefix}_dl_csv_{key_no}")
    with c2:
        if FPDF_AVAILABLE:
            pdf_bytes = build_pdf(pdf_title, record, signature_images)
            st.download_button("⬇️ PDF Receipt", data=pdf_bytes, file_name=f"{prefix}_{key_no}.pdf",
                                mime="application/pdf", use_container_width=True, key=f"{prefix}_dl_pdf_{key_no}")
        else:
            st.caption("PDF export needs `fpdf2` installed.")
    with c3:
        if OPENPYXL_AVAILABLE:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                pd.DataFrame([record]).to_excel(writer, index=False, sheet_name="Record")
            st.download_button("⬇️ Excel", data=buf.getvalue(), file_name=f"{prefix}_{key_no}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True, key=f"{prefix}_dl_xlsx_{key_no}")
        else:
            st.caption("Excel export needs `openpyxl` installed.")


def records_viewer(file_name: str, prefix: str):
    with st.expander(f"📋 View All Saved {prefix.replace('_', ' ').title()} Records"):
        if os.path.exists(file_name):
            df_all = pd.read_csv(file_name)
            st.dataframe(df_all, use_container_width=True)
            e1, e2 = st.columns(2)
            with e1:
                st.download_button("⬇️ Download all as CSV", data=df_all.to_csv(index=False).encode("utf-8"),
                                    file_name=f"all_{prefix}.csv", mime="text/csv",
                                    use_container_width=True, key=f"{prefix}_all_csv")
            with e2:
                if OPENPYXL_AVAILABLE:
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        df_all.to_excel(writer, index=False, sheet_name="Records")
                    st.download_button("⬇️ Download all as Excel", data=buf.getvalue(),
                                        file_name=f"all_{prefix}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True, key=f"{prefix}_all_xlsx")
                else:
                    st.caption("Excel export needs `openpyxl` installed.")
        else:
            st.info("No records saved yet.")


# ---------------------------------------------------------------------------
# TABS
# ---------------------------------------------------------------------------
tab_withdraw, tab_transfer, tab_deposit, tab_reports = st.tabs(
    ["💵 Cash Withdrawal", "🔁 Account to Account Transfer", "💰 Deposit Form", "📊 Reports"]
)

# ===========================================================================
# TAB 1 — CASH WITHDRAWAL
# ===========================================================================
with tab_withdraw:
    st.subheader("CASH WITHDRAWAL FORM")
    st.caption("Unka Baasiin Ittiin Ajajamu / ገንዘብ ወጪ ማድረጊያ ቅፅ")

    dcol1, dcol2 = st.columns(2)
    with dcol1:
        st.button("🎬 Load Demo Data", key="wd_demo_btn", use_container_width=True,
                   on_click=load_demo_data, args=({
                       "wd_branch_choice": "Gerji", "wd_branch": "Gerji",
                       "wd_acct_type": "Demand Account (Herrega Socho'aa)",
                       "wd_full_name": "Jorgu Adulu Bilisaa", "wd_acct_no": "1907478600007",
                       "wd_amount": 20000, "wd_words": "twenty thousand",
                       "wd_x200": 100, "wd_x100": 0, "wd_x50": 0, "wd_x10": 0, "wd_x5": 0, "wd_x1": 0, "wd_extra": 0.0,
                       "wd_c1_name": "Jorgu Adulu Bilisaa", "wd_c1_phone": "+251911234567",
                       "wd_c2_name": "", "wd_c2_phone": "",
                   },))
    with dcol2:
        st.button("🧹 Clear Form", key="wd_clear_btn", use_container_width=True,
                   on_click=clear_form_fields, args=([
                       "wd_branch_choice", "wd_branch", "wd_acct_type", "wd_other", "wd_full_name", "wd_acct_no",
                       "wd_amount", "wd_words", "wd_x200", "wd_x100", "wd_x50", "wd_x10", "wd_x5", "wd_x1", "wd_extra",
                       "wd_c1_name", "wd_c1_phone", "wd_c1_typed", "wd_c2_name", "wd_c2_phone", "wd_c2_typed",
                   ],))

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    wd_date = date_field("wd_date")
    wd_branch = branch_selector("wd")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    wd_acct_type, wd_other = account_type_block("wd")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        wd_full_name = st.text_input("Full Name / Maqaa Guutuu (ሙሉ ስም)", key="wd_full_name")
    with c2:
        wd_acct_no = st.text_input("A/C No. / Lakk. Herrega (የሂሳብ ቁጥር)", key="wd_acct_no")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    wd_amount = st.number_input("Amount in Number / Qarshii (በቁጥር)", min_value=0, step=1, format="%d", key="wd_amount")
    wd_words = st.text_input("Amount in Words / Hanga Qarshii Jechaan (በፊደል)", key="wd_words")
    st.markdown('<div class="ai-box">🤖 <b>AI Assist</b></div>', unsafe_allow_html=True)
    st.button("✨ Auto-write Amount in Words", key="wd_words_btn", on_click=ai_auto_words, args=("wd_amount", "wd_words"))
    st.markdown("**Denomination Needed / Jijjiirraa Qarshii Barbaaddani**")
    wd_denom_vals, wd_cents, wd_total = denomination_block("wd", include_ones=True, extra_label="CENTS", amount_key="wd_amount")
    if wd_amount > 0 and wd_total != wd_amount:
        st.warning(f"⚠️ Denomination total ({wd_total:,.2f}) doesn't match Amount ({wd_amount:,}).")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown("**Customer Signatures**")
    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("① Customer 1")
        wd_c1_name = st.text_input("Name of Customer (1)", key="wd_c1_name")
        wd_c1_sig = signature_pad("wd_c1", "Signature (1)")
        wd_c1_phone = st.text_input("Tel/Mob (1)", key="wd_c1_phone")
        st.button("✨ Format Phone Number", key="wd_c1_phone_btn", on_click=ai_format_phone, args=("wd_c1_phone",))
    with cc2:
        st.markdown("② Customer 2")
        wd_c2_name = st.text_input("Name of Customer (2)", key="wd_c2_name")
        wd_c2_sig = signature_pad("wd_c2", "Signature (2)")
        wd_c2_phone = st.text_input("Tel/Mob (2)", key="wd_c2_phone")
        st.button("✨ Format Phone Number", key="wd_c2_phone_btn", on_click=ai_format_phone, args=("wd_c2_phone",))
    st.markdown('</div>', unsafe_allow_html=True)
    st.caption("Notice: Passbook must accompany with this form.")

    if st.button("🤖 Run Smart Check", key="wd_smart_check", use_container_width=True):
        issues = ai_smart_check({
            "required": {"Branch": wd_branch, "Full Name": wd_full_name, "Account No": wd_acct_no, "Amount": wd_amount},
            "phone": wd_c1_phone, "account_no": wd_acct_no,
            "denom_total": wd_total, "amount": wd_amount,
        })
        if issues:
            st.warning("Smart Check found some issues:\n\n" + "\n".join(f"- {i}" for i in issues))
        else:
            st.success("Smart Check passed — form looks complete and consistent.")

    if st.button("✅ Submit Withdrawal Form", key="wd_submit", use_container_width=True):
        if not wd_branch or not wd_full_name or not wd_acct_no or wd_amount <= 0:
            st.error("Please fill in Branch, Full Name, Account Number, and Amount.")
        else:
            if check_duplicate("withdrawal_records.csv", wd_acct_no, wd_amount):
                st.warning("🤖 Possible duplicate: a withdrawal with this account number and amount was submitted in the last 10 minutes. Please double-check before proceeding.")
            sig1_bytes = wd_c1_sig if isinstance(wd_c1_sig, (bytes, bytearray)) else None
            sig2_bytes = wd_c2_sig if isinstance(wd_c2_sig, (bytes, bytearray)) else None
            record = {
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Date": wd_date.strftime("%d/%m/%Y"),
                "Branch": wd_branch,
                "Account Type": wd_other if wd_acct_type == "Other (Specify)" else wd_acct_type,
                "Full Name": wd_full_name,
                "Account No": wd_acct_no,
                "Amount in Words": wd_words,
                "Amount in Number": wd_amount,
                "X200": wd_denom_vals[200], "X100": wd_denom_vals[100], "X50": wd_denom_vals[50],
                "X10": wd_denom_vals[10], "X5": wd_denom_vals[5], "X1": wd_denom_vals[1], "Cents": wd_cents,
                "Denomination Total": wd_total,
                "Customer1 Name": wd_c1_name,
                "Customer1 Signature": "Drawn (see PDF)" if sig1_bytes else (wd_c1_sig or ""),
                "Customer1 Phone": wd_c1_phone,
                "Customer2 Name": wd_c2_name,
                "Customer2 Signature": "Drawn (see PDF)" if sig2_bytes else (wd_c2_sig or ""),
                "Customer2 Phone": wd_c2_phone,
            }
            save_record("withdrawal_records.csv", record)
            st.success("Withdrawal form submitted and saved!")
            st.table(pd.DataFrame([record]).T.rename(columns={0: "Value"}))
            export_buttons("withdrawal", record, "Cash Withdrawal Receipt",
                            {"Customer 1 Signature": sig1_bytes, "Customer 2 Signature": sig2_bytes},
                            wd_acct_no)

    records_viewer("withdrawal_records.csv", "cash_withdrawal")

# ===========================================================================
# TAB 2 — ACCOUNT TO ACCOUNT TRANSFER
# ===========================================================================
with tab_transfer:
    st.subheader("ACCOUNT TO ACCOUNT TRANSFER")
    st.caption("Unka Herregaa Gara Herregaatti Maallaqa Daddabarsan")

    dcol1, dcol2 = st.columns(2)
    with dcol1:
        st.button("🎬 Load Demo Data", key="tr_demo_btn", use_container_width=True,
                   on_click=load_demo_data, args=({
                       "tr_branch_choice": "Bole", "tr_branch": "Bole",
                       "tr_acct_type": "Saving Account (Herrega Qusannoo)",
                       "tr_full_name": "Tamirat Worku", "tr_acct_no": "18645705000041",
                       "tr_amount": 20000, "tr_words": "twenty thousand",
                       "tr_ben_name": "Almaz Kebede", "tr_ben_acct_no": "10098765432101",
                       "tr_ben_branch": "Piassa", "tr_narrative": "Family Support",
                       "tr_app_name": "Tamirat Worku", "tr_city": "Addis Ababa", "tr_subcity": "Bole",
                       "tr_woreda": "03", "tr_house_no": "245", "tr_phone": "+251922334455",
                   },))
    with dcol2:
        st.button("🧹 Clear Form", key="tr_clear_btn", use_container_width=True,
                   on_click=clear_form_fields, args=([
                       "tr_branch_choice", "tr_branch", "tr_acct_type", "tr_other", "tr_full_name", "tr_acct_no",
                       "tr_amount", "tr_words", "tr_x200", "tr_x100", "tr_x50", "tr_x10", "tr_x5",
                       "tr_ben_name", "tr_ben_acct_no", "tr_ben_branch", "tr_narrative_choice", "tr_narrative",
                       "tr_app_name", "tr_app_typed", "tr_city", "tr_subcity", "tr_woreda", "tr_house_no", "tr_phone",
                   ],))

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    tr_date = date_field("tr_date")
    tr_branch = branch_selector("tr")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    tr_acct_type, tr_other = account_type_block("tr")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown("**Debit Account (Herrega irraa Hir'ifamu)**")
    c1, c2 = st.columns(2)
    with c1:
        tr_full_name = st.text_input("Full Name / Maqaa Guutuu", key="tr_full_name")
    with c2:
        tr_acct_no = st.text_input("A/C No. / Lakk. Herrega", key="tr_acct_no")
    tr_amount = st.number_input("Amount in Number / Qarshii", min_value=0, step=1, format="%d", key="tr_amount")
    tr_words = st.text_input("Amount in Words / Hanga Qarshii Jechaan", key="tr_words")
    st.markdown('<div class="ai-box">🤖 <b>AI Assist</b></div>', unsafe_allow_html=True)
    st.button("✨ Auto-write Amount in Words", key="tr_words_btn", on_click=ai_auto_words, args=("tr_amount", "tr_words"))
    st.markdown("**Denomination (optional add-on, not on the original paper form)**")
    tr_denom_vals, _, tr_total = denomination_block("tr", include_ones=False, extra_label=None, amount_key="tr_amount")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown("**Beneficiary's Account (Herrega galii ta'uuf)**")
    c3, c4 = st.columns(2)
    with c3:
        tr_ben_name = st.text_input("Beneficiary Full Name", key="tr_ben_name")
    with c4:
        tr_ben_acct_no = st.text_input("Beneficiary A/C No.", key="tr_ben_acct_no")
    tr_ben_branch = st.text_input("Branch Where Beneficiary Account is Maintained", key="tr_ben_branch")
    narrative_choice = st.selectbox("🤖 Quick-pick narrative (or choose Other and type below)",
                                     COMMON_NARRATIVES, key="tr_narrative_choice")
    if narrative_choice != "Other":
        st.session_state["tr_narrative"] = narrative_choice
    tr_narrative = st.text_input("Payment Narrative / Sababa Kaffaltii", key="tr_narrative")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown("**Applicant & Address of Depositor**")
    tr_app_name = st.text_input("Applicant Name / Maqaa Iyyataa", key="tr_app_name")
    tr_app_sig = signature_pad("tr_app", "Applicant Signature / Mallattoo")
    ac1, ac2 = st.columns(2)
    with ac1:
        tr_city = st.text_input("City / Magaalaa", key="tr_city")
    with ac2:
        tr_subcity = st.text_input("Sub-city / Kutaa Magaalaa", key="tr_subcity")
    ac3, ac4 = st.columns(2)
    with ac3:
        tr_woreda = st.text_input("Woreda / Aanaa", key="tr_woreda")
    with ac4:
        tr_house_no = st.text_input("House No. / Lakk. Mana", key="tr_house_no")
    tr_phone = st.text_input("Tel/Mob / Bilbila", key="tr_phone")
    st.button("✨ Format Phone Number", key="tr_phone_btn", on_click=ai_format_phone, args=("tr_phone",))
    st.markdown('</div>', unsafe_allow_html=True)
    st.caption("Notice: This account to account transfer shall take the place of the account owner; ID card may be required.")

    if st.button("🤖 Run Smart Check", key="tr_smart_check", use_container_width=True):
        issues = ai_smart_check({
            "required": {"Branch": tr_branch, "Full Name": tr_full_name, "Account No": tr_acct_no,
                          "Amount": tr_amount, "Beneficiary Account No": tr_ben_acct_no},
            "phone": tr_phone, "account_no": tr_acct_no,
        })
        if issues:
            st.warning("Smart Check found some issues:\n\n" + "\n".join(f"- {i}" for i in issues))
        else:
            st.success("Smart Check passed — form looks complete and consistent.")

    if st.button("✅ Submit Transfer Form", key="tr_submit", use_container_width=True):
        if not tr_branch or not tr_full_name or not tr_acct_no or tr_amount <= 0 or not tr_ben_acct_no:
            st.error("Please fill in Branch, Full Name, Debit Account No., Amount, and Beneficiary Account No.")
        else:
            if check_duplicate("transfer_records.csv", tr_acct_no, tr_amount):
                st.warning("🤖 Possible duplicate: a transfer with this account number and amount was submitted in the last 10 minutes. Please double-check before proceeding.")
            sig_bytes = tr_app_sig if isinstance(tr_app_sig, (bytes, bytearray)) else None
            record = {
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Date": tr_date.strftime("%d/%m/%Y"),
                "Branch": tr_branch,
                "Account Type": tr_other if tr_acct_type == "Other (Specify)" else tr_acct_type,
                "Debit Full Name": tr_full_name,
                "Debit Account No": tr_acct_no,
                "Amount in Words": tr_words,
                "Amount in Number": tr_amount,
                "X200": tr_denom_vals.get(200, 0), "X100": tr_denom_vals.get(100, 0),
                "X50": tr_denom_vals.get(50, 0), "X10": tr_denom_vals.get(10, 0), "X5": tr_denom_vals.get(5, 0),
                "Denomination Total": tr_total,
                "Beneficiary Full Name": tr_ben_name,
                "Beneficiary Account No": tr_ben_acct_no,
                "Beneficiary Branch": tr_ben_branch,
                "Payment Narrative": tr_narrative,
                "Applicant Name": tr_app_name,
                "Applicant Signature": "Drawn (see PDF)" if sig_bytes else (tr_app_sig or ""),
                "City": tr_city, "Sub-city": tr_subcity, "Woreda": tr_woreda,
                "House No": tr_house_no, "Phone": tr_phone,
            }
            save_record("transfer_records.csv", record)
            st.success("Transfer form submitted and saved!")
            st.table(pd.DataFrame([record]).T.rename(columns={0: "Value"}))
            export_buttons("transfer", record, "Account to Account Transfer Receipt",
                            {"Applicant Signature": sig_bytes}, tr_acct_no)

    records_viewer("transfer_records.csv", "account_transfer")

# ===========================================================================
# TAB 3 — DEPOSIT FORM
# ===========================================================================
with tab_deposit:
    st.subheader("DEPOSIT FORM")
    st.caption("Unka Galiin Ittiin Ajajamu / ገቢ ማድረጊያ ቅፅ")

    dcol1, dcol2 = st.columns(2)
    with dcol1:
        st.button("🎬 Load Demo Data", key="dp_demo_btn", use_container_width=True,
                   on_click=load_demo_data, args=({
                       "dp_branch_choice": "Jimma", "dp_branch": "Jimma",
                       "dp_acct_type": "Saving Account (Herrega Qusannoo)",
                       "dp_full_name": "Chaltu Bekele", "dp_acct_no": "1122334455667",
                       "dp_x200": 50, "dp_x100": 0, "dp_x50": 0, "dp_x10": 0, "dp_x5": 0, "dp_x1": 0, "dp_extra": 0.0,
                       "dp_words": "ten thousand", "dp_source": "Business income",
                       "dp_dep_name": "Chaltu Bekele", "dp_city": "Jimma", "dp_subcity": "Merkato",
                       "dp_woreda": "01", "dp_house_no": "12", "dp_phone": "+251933445566",
                   },))
    with dcol2:
        st.button("🧹 Clear Form", key="dp_clear_btn", use_container_width=True,
                   on_click=clear_form_fields, args=([
                       "dp_branch_choice", "dp_branch", "dp_acct_type", "dp_other", "dp_full_name", "dp_acct_no",
                       "dp_cheque_branch", "dp_cheque_no", "dp_drawer_acct",
                       "dp_x200", "dp_x100", "dp_x50", "dp_x10", "dp_x5", "dp_x1", "dp_extra",
                       "dp_words", "dp_source", "dp_dep_name", "dp_dep_typed",
                       "dp_city", "dp_subcity", "dp_woreda", "dp_house_no", "dp_phone",
                   ],))

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    dp_date = date_field("dp_date")
    dp_branch = branch_selector("dp")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    dp_acct_type, dp_other = account_type_block("dp")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        dp_full_name = st.text_input("Full Name / Maqaa Guutuu", key="dp_full_name")
    with c2:
        dp_acct_no = st.text_input("A/C No. / Lakk. Herrega", key="dp_acct_no")
    st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("🧾 Cheque Deposit (optional — only if depositing a cheque)"):
        dp_cheque_branch = st.text_input("Cheque Branch / Damee", key="dp_cheque_branch")
        dp_cheque_no = st.text_input("Cheque No. / Lakk. Cheekii", key="dp_cheque_no")
        dp_drawer_acct = st.text_input("Drawer A/C No. / Lakk. Herrega", key="dp_drawer_acct")

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown("**Cash Deposit — Denomination / Jijjiirraa Qarshii**")
    dp_denom_vals, dp_coins, dp_total = denomination_block("dp", include_ones=True, extra_label="Coins", amount_key=None)
    st.info(f"Amount (auto-computed from denomination) = **{dp_total:,.2f}**")
    dp_words = st.text_input("Amount in Words / Hanga Qarshii Jechaan", key="dp_words")
    st.button("✨ Auto-write Amount in Words (from denomination total)", key="dp_words_btn",
              on_click=lambda: st.session_state.update(
                  {"dp_words": number_to_words(int(dp_total))} if dp_total > 0 else {}))
    dp_source = st.text_input("Source of Proceeds / Madda Maallaqaa", key="dp_source")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown("**Depositor**")
    dp_dep_name = st.text_input("Name of Depositor / Maqaa Nama Galchee", key="dp_dep_name")
    dp_dep_sig = signature_pad("dp_dep", "Signature / Mallattoo")
    ac1, ac2 = st.columns(2)
    with ac1:
        dp_city = st.text_input("City / Magaalaa", key="dp_city")
    with ac2:
        dp_subcity = st.text_input("Sub-city / Kutaa Magaalaa", key="dp_subcity")
    ac3, ac4 = st.columns(2)
    with ac3:
        dp_woreda = st.text_input("Woreda / Aanaa", key="dp_woreda")
    with ac4:
        dp_house_no = st.text_input("House No. / Lakk. Mana", key="dp_house_no")
    dp_phone = st.text_input("Tel/Mob / Bilbila", key="dp_phone")
    st.button("✨ Format Phone Number", key="dp_phone_btn", on_click=ai_format_phone, args=("dp_phone",))
    st.markdown('</div>', unsafe_allow_html=True)
    st.caption("Notice: This Deposit Form cannot be used as a Receipt. Cheque deposits are subject to clearance.")

    if st.button("🤖 Run Smart Check", key="dp_smart_check", use_container_width=True):
        issues = ai_smart_check({
            "required": {"Branch": dp_branch, "Full Name": dp_full_name, "Account No": dp_acct_no,
                          "Denomination Total": dp_total},
            "phone": dp_phone, "account_no": dp_acct_no,
        })
        if issues:
            st.warning("Smart Check found some issues:\n\n" + "\n".join(f"- {i}" for i in issues))
        else:
            st.success("Smart Check passed — form looks complete and consistent.")

    if st.button("✅ Submit Deposit Form", key="dp_submit", use_container_width=True):
        if not dp_branch or not dp_full_name or not dp_acct_no or dp_total <= 0:
            st.error("Please fill in Branch, Full Name, Account Number, and at least one denomination amount.")
        else:
            if check_duplicate("deposit_records.csv", dp_acct_no, dp_total):
                st.warning("🤖 Possible duplicate: a deposit with this account number and amount was submitted in the last 10 minutes. Please double-check before proceeding.")
            sig_bytes = dp_dep_sig if isinstance(dp_dep_sig, (bytes, bytearray)) else None
            record = {
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Date": dp_date.strftime("%d/%m/%Y"),
                "Branch": dp_branch,
                "Account Type": dp_other if dp_acct_type == "Other (Specify)" else dp_acct_type,
                "Full Name": dp_full_name,
                "Account No": dp_acct_no,
                "Cheque Branch": dp_cheque_branch, "Cheque No": dp_cheque_no, "Drawer Account No": dp_drawer_acct,
                "X200": dp_denom_vals[200], "X100": dp_denom_vals[100], "X50": dp_denom_vals[50],
                "X10": dp_denom_vals[10], "X5": dp_denom_vals[5], "X1": dp_denom_vals[1], "Coins": dp_coins,
                "Amount (Denomination Total)": dp_total,
                "Amount in Words": dp_words,
                "Source of Proceeds": dp_source,
                "Depositor Name": dp_dep_name,
                "Depositor Signature": "Drawn (see PDF)" if sig_bytes else (dp_dep_sig or ""),
                "City": dp_city, "Sub-city": dp_subcity, "Woreda": dp_woreda,
                "House No": dp_house_no, "Phone": dp_phone,
            }
            save_record("deposit_records.csv", record)
            st.success("Deposit form submitted and saved!")
            st.table(pd.DataFrame([record]).T.rename(columns={0: "Value"}))
            export_buttons("deposit", record, "Deposit Form Receipt",
                            {"Depositor Signature": sig_bytes}, dp_acct_no)

    records_viewer("deposit_records.csv", "deposit")

# ===========================================================================
# TAB 4 — REPORTS (one combined system across all three forms)
# ===========================================================================
with tab_reports:
    st.subheader("📊 COMBINED REPORT — ALL FORMS")
    st.caption("One system, one report — Cash Withdrawal + Account Transfer + Deposit")

    sources = {
        "Cash Withdrawal": ("withdrawal_records.csv", "Amount in Number"),
        "Account Transfer": ("transfer_records.csv", "Amount in Number"),
        "Deposit": ("deposit_records.csv", "Amount (Denomination Total)"),
    }

    combined_rows = []
    for form_type, (fname, amt_col) in sources.items():
        if os.path.exists(fname):
            df = pd.read_csv(fname)
            for _, row in df.iterrows():
                combined_rows.append({
                    "Form Type": form_type,
                    "Timestamp": row.get("Timestamp", ""),
                    "Date": row.get("Date", ""),
                    "Branch": row.get("Branch", ""),
                    "Full Name": row.get("Full Name", row.get("Debit Full Name", "")),
                    "Account No": row.get("Account No", row.get("Debit Account No", "")),
                    "Amount": row.get(amt_col, 0),
                })

    if not combined_rows:
        st.info("No records submitted yet across any of the three forms. Once tellers start submitting, a combined report will appear here.")
    else:
        df_all = pd.DataFrame(combined_rows)
        df_all["Amount"] = pd.to_numeric(df_all["Amount"], errors="coerce").fillna(0)

        # ---- Summary metrics ----
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Transactions", len(df_all))
        m2.metric("Total Withdrawals", f"{df_all[df_all['Form Type']=='Cash Withdrawal']['Amount'].sum():,.0f}")
        m3.metric("Total Transfers", f"{df_all[df_all['Form Type']=='Account Transfer']['Amount'].sum():,.0f}")
        m4.metric("Total Deposits", f"{df_all[df_all['Form Type']=='Deposit']['Amount'].sum():,.0f}")

        st.markdown("---")

        # ---- Filters ----
        f1, f2 = st.columns(2)
        with f1:
            type_filter = st.multiselect("Filter by Form Type", options=list(sources.keys()),
                                          default=list(sources.keys()), key="rep_type_filter")
        with f2:
            branch_options = sorted(df_all["Branch"].dropna().unique().tolist())
            branch_filter = st.multiselect("Filter by Branch", options=branch_options,
                                            default=branch_options, key="rep_branch_filter")

        filtered = df_all[df_all["Form Type"].isin(type_filter) & df_all["Branch"].isin(branch_filter)]

        st.markdown(f"**Showing {len(filtered)} of {len(df_all)} transactions**")
        st.dataframe(filtered.sort_values("Timestamp", ascending=False), use_container_width=True)

        # ---- Chart: total amount by form type ----
        chart_data = filtered.groupby("Form Type")["Amount"].sum()
        if not chart_data.empty:
            st.markdown("**Total Amount by Form Type**")
            st.bar_chart(chart_data)

        # ---- Chart: total amount by branch ----
        branch_chart = filtered.groupby("Branch")["Amount"].sum().sort_values(ascending=False)
        if not branch_chart.empty:
            st.markdown("**Total Amount by Branch**")
            st.bar_chart(branch_chart)

        st.markdown("---")

        # ---- Combined export ----
        e1, e2 = st.columns(2)
        with e1:
            st.download_button("⬇️ Download Combined Report (CSV)",
                                data=filtered.to_csv(index=False).encode("utf-8"),
                                file_name="combined_report.csv", mime="text/csv",
                                use_container_width=True, key="rep_dl_csv")
        with e2:
            if OPENPYXL_AVAILABLE:
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    filtered.to_excel(writer, index=False, sheet_name="Combined Report")
                    for form_type, (fname, _) in sources.items():
                        if os.path.exists(fname):
                            pd.read_csv(fname).to_excel(writer, index=False, sheet_name=form_type[:31])
                st.download_button("⬇️ Download Combined Report (Excel, all sheets)",
                                    data=buf.getvalue(), file_name="combined_report.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True, key="rep_dl_xlsx")
            else:
                st.caption("Excel export needs `openpyxl` installed.")
